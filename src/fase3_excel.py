"""
FASE 3 — Template Excel para el cliente
=========================================
Herramientas: openpyxl, pandas
Input:  data/processed/portfolio_data.csv  (generado en Fase 1)
Output: excel/cartera_template.xlsx

Hojas generadas:
    1. Mi Cartera     — ingreso de activos y cálculo automático de P&L
    2. Precios        — precios históricos traídos del CSV
    3. Resumen        — KPIs consolidados de la cartera

Uso:
    python src/fase3_excel.py
"""

import sys
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

sys.stdout.reconfigure(encoding="utf-8")


# ─────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────

INPUT_CSV    = "data/processed/portfolio_data.csv"
OUTPUT_EXCEL = "excel/cartera_template.xlsx"

# Paleta de colores
AZUL_OSCURO  = "0C447C"
AZUL_MED     = "185FA5"
AZUL_CLARO   = "D6E8F7"
VERDE_OSCURO = "1E6B34"
VERDE_CLARO  = "D6F0DE"
ROJO_OSCURO  = "8B1A1A"
ROJO_CLARO   = "FADADD"
GRIS_HEADER  = "F1F1F1"
GRIS_BORDE   = "CCCCCC"
BLANCO       = "FFFFFF"
AMARILLO     = "FFF9C4"

# Cartera demo — mismos tickers que Fase 1
CARTERA_DEMO = [
    ("AAPL",  "Apple Inc.",        150,  145.20),
    ("MSFT",  "Microsoft",          80,  290.50),
    ("GOOGL", "Alphabet",           40,  138.75),
    ("BRK-B", "Berkshire Hathaway", 60,  278.30),
    ("JPM",   "JPMorgan Chase",    100,  148.60),
    ("GLD",   "SPDR Gold ETF",      50,  168.40),
    ("SPY",   "S&P 500 ETF",        30,  430.00),
]


# ─────────────────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────────────────

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def border_thin():
    side = Side(style="thin", color=GRIS_BORDE)
    return Border(left=side, right=side, top=side, bottom=side)

def border_medium():
    side = Side(style="medium", color=AZUL_OSCURO)
    return Border(left=side, right=side, top=side, bottom=side)

def estilo_header(cell, color_fondo=AZUL_OSCURO, color_texto=BLANCO, size=11):
    cell.font      = Font(bold=True, color=color_texto, size=size, name="Calibri")
    cell.fill      = fill(color_fondo)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border_thin()

def estilo_celda(cell, bold=False, color_fondo=BLANCO, color_texto="000000",
                 alineacion="left", formato=None):
    cell.font      = Font(bold=bold, color=color_texto, size=10, name="Calibri")
    cell.fill      = fill(color_fondo)
    cell.alignment = Alignment(horizontal=alineacion, vertical="center")
    cell.border    = border_thin()
    if formato:
        cell.number_format = formato

def ajustar_columnas(ws, anchos: dict):
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

def titulo_hoja(ws, texto, subtexto=None):
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = texto
    c.font      = Font(bold=True, size=16, color=BLANCO, name="Calibri")
    c.fill      = fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36

    if subtexto:
        ws.merge_cells("A2:I2")
        c2 = ws["A2"]
        c2.value     = subtexto
        c2.font      = Font(size=10, color=AZUL_MED, italic=True, name="Calibri")
        c2.fill      = fill(AZUL_CLARO)
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20


# ─────────────────────────────────────────────────────────
# HOJA 1 — MI CARTERA
# ─────────────────────────────────────────────────────────

def crear_hoja_cartera(wb: Workbook, precios: pd.DataFrame):
    """
    Hoja principal editable. El cliente ingresa:
        - Ticker
        - Nombre del activo
        - Cantidad de unidades
        - Precio de compra (USD)

    El template calcula automáticamente:
        - Precio actual (último del CSV)
        - Valor de compra total
        - Valor actual total
        - P&L en USD
        - P&L en %
        - Peso en la cartera (%)
    """
    ws = wb.create_sheet("Mi Cartera")
    wb.active = ws

    titulo_hoja(
        ws,
        "MI CARTERA DE INVERSIONES",
        "Completá las columnas en celeste  ·  El resto se calcula automáticamente"
    )

    # ── Encabezados ──────────────────────────────────────
    fila_header = 4
    headers = [
        ("A", "Ticker",            12),
        ("B", "Nombre del activo", 22),
        ("C", "Unidades",          12),
        ("D", "Precio compra (USD)",15),
        ("E", "Precio actual (USD)",15),
        ("F", "Valor compra (USD)", 16),
        ("G", "Valor actual (USD)", 16),
        ("H", "P&L (USD)",          13),
        ("I", "P&L (%)",            10),
        ("J", "Peso (%)",           10),
    ]

    for col, nombre, ancho in headers:
        c = ws[f"{col}{fila_header}"]
        c.value = nombre
        # Columnas editables por el cliente en celeste
        if col in ["A", "B", "C", "D"]:
            estilo_header(c, color_fondo=AZUL_MED, color_texto=BLANCO)
        else:
            estilo_header(c, color_fondo=AZUL_OSCURO, color_texto=BLANCO)
        ws.column_dimensions[col].width = ancho

    ws.row_dimensions[fila_header].height = 36

    # ── Datos demo ───────────────────────────────────────
    # Obtener últimos precios del CSV
    ultimo_precio = precios.iloc[-1].to_dict()

    primera_data = fila_header + 1
    for i, (ticker, nombre, unidades, precio_compra) in enumerate(CARTERA_DEMO):
        fila = primera_data + i
        color_fila = GRIS_HEADER if i % 2 == 0 else BLANCO

        precio_actual = round(ultimo_precio.get(ticker, 0), 2)

        # Columnas editables (celeste claro)
        for col, val, fmt in [
            ("A", ticker,        "@"),
            ("B", nombre,        "@"),
            ("C", unidades,      "#,##0"),
            ("D", precio_compra, "#,##0.00"),
        ]:
            c = ws[f"{col}{fila}"]
            c.value = val
            estilo_celda(c, color_fondo=AZUL_CLARO, alineacion="center" if col in ["A","C","D"] else "left", formato=fmt)

        # Precio actual (traído del CSV, no editable)
        c = ws[f"E{fila}"]
        c.value = precio_actual
        estilo_celda(c, color_fondo=color_fila, alineacion="center", formato="#,##0.00")

        # Valor de compra = Unidades × Precio compra
        c = ws[f"F{fila}"]
        c.value = f"=C{fila}*D{fila}"
        estilo_celda(c, color_fondo=color_fila, alineacion="center", formato="#,##0.00")

        # Valor actual = Unidades × Precio actual
        c = ws[f"G{fila}"]
        c.value = f"=C{fila}*E{fila}"
        estilo_celda(c, color_fondo=color_fila, alineacion="center", formato="#,##0.00")

        # P&L USD = Valor actual - Valor compra
        c = ws[f"H{fila}"]
        c.value = f"=G{fila}-F{fila}"
        estilo_celda(c, color_fondo=color_fila, alineacion="center", formato='+#,##0.00;-#,##0.00')

        # P&L % = (Valor actual / Valor compra - 1) × 100
        c = ws[f"I{fila}"]
        c.value = f"=(G{fila}/F{fila}-1)*100"
        estilo_celda(c, color_fondo=color_fila, alineacion="center", formato='+0.00%;-0.00%')

        # Peso % (se completa después con fórmula que usa total)
        c = ws[f"J{fila}"]
        c.value = f"=G{fila}/G${primera_data + len(CARTERA_DEMO)}*100"
        estilo_celda(c, color_fondo=color_fila, alineacion="center", formato='0.00"%"')

    # ── Fila de totales ───────────────────────────────────
    fila_total = primera_data + len(CARTERA_DEMO)
    ws.row_dimensions[fila_total].height = 24

    ws.merge_cells(f"A{fila_total}:B{fila_total}")
    c = ws[f"A{fila_total}"]
    c.value = "TOTAL CARTERA"
    estilo_celda(c, bold=True, color_fondo=AZUL_OSCURO, color_texto=BLANCO, alineacion="center")

    for col in ["C", "D", "E"]:
        c = ws[f"{col}{fila_total}"]
        estilo_celda(c, color_fondo=AZUL_OSCURO)

    for col, formula in [
        ("F", f"=SUM(F{primera_data}:F{fila_total-1})"),
        ("G", f"=SUM(G{primera_data}:G{fila_total-1})"),
        ("H", f"=SUM(H{primera_data}:H{fila_total-1})"),
        ("I", f"=(G{fila_total}/F{fila_total}-1)*100"),
        ("J", f"=SUM(J{primera_data}:J{fila_total-1})"),
    ]:
        c = ws[f"{col}{fila_total}"]
        c.value = formula
        estilo_celda(c, bold=True, color_fondo=AZUL_OSCURO, color_texto=BLANCO,
                     alineacion="center",
                     formato="#,##0.00" if col in ["F","G","H"] else '0.00"%"')

    # ── Nota al pie ──────────────────────────────────────
    fila_nota = fila_total + 2
    ws.merge_cells(f"A{fila_nota}:J{fila_nota}")
    c = ws[f"A{fila_nota}"]
    c.value = "ℹ  Columnas en celeste son editables. Precio actual corresponde al último dato disponible del CSV generado por el pipeline Python."
    c.font      = Font(size=9, italic=True, color=AZUL_MED, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Congelar paneles en la fila de datos
    ws.freeze_panes = f"A{primera_data}"

    print("  ✓ Hoja 'Mi Cartera' creada")
    return primera_data, fila_total


# ─────────────────────────────────────────────────────────
# HOJA 2 — PRECIOS
# ─────────────────────────────────────────────────────────

def crear_hoja_precios(wb: Workbook, precios: pd.DataFrame):
    """
    Muestra los últimos 60 días de precios de cierre por activo.
    Referencia para el cliente y fuente de datos para fórmulas.
    """
    ws = wb.create_sheet("Precios")

    titulo_hoja(
        ws,
        "PRECIOS HISTÓRICOS DE CIERRE",
        "Últimos 60 días hábiles  ·  Fuente: Yahoo Finance vía yfinance"
    )

    # Últimos 60 días
    precios_recientes = precios.tail(60).copy()
    tickers = precios_recientes.columns.tolist()

    # Header fecha
    fila_header = 4
    c = ws[f"A{fila_header}"]
    c.value = "Fecha"
    estilo_header(c)
    ws.column_dimensions["A"].width = 14

    # Headers tickers
    for j, ticker in enumerate(tickers):
        col = get_column_letter(j + 2)
        c = ws[f"{col}{fila_header}"]
        c.value = ticker
        estilo_header(c)
        ws.column_dimensions[col].width = 13

    ws.row_dimensions[fila_header].height = 28

    # Datos
    for i, (fecha, row) in enumerate(precios_recientes.iterrows()):
        fila = fila_header + 1 + i
        color_fila = GRIS_HEADER if i % 2 == 0 else BLANCO

        c = ws[f"A{fila}"]
        c.value = fecha.strftime("%Y-%m-%d")
        estilo_celda(c, color_fondo=color_fila, alineacion="center")

        for j, ticker in enumerate(tickers):
            col = get_column_letter(j + 2)
            c = ws[f"{col}{fila}"]
            c.value = round(row[ticker], 2)
            estilo_celda(c, color_fondo=color_fila, alineacion="center", formato="#,##0.00")

    ws.freeze_panes = f"A{fila_header + 1}"

    print("  ✓ Hoja 'Precios' creada")


# ─────────────────────────────────────────────────────────
# HOJA 3 — RESUMEN
# ─────────────────────────────────────────────────────────

def crear_hoja_resumen(wb: Workbook, precios: pd.DataFrame, primera_data: int, fila_total: int):
    """
    KPIs ejecutivos consolidados de la cartera.
    Referencia directamente a los cálculos de 'Mi Cartera'.
    """
    ws = wb.create_sheet("Resumen")

    titulo_hoja(
        ws,
        "RESUMEN EJECUTIVO DE LA CARTERA",
        "KPIs consolidados  ·  Se actualiza automáticamente al modificar 'Mi Cartera'"
    )

    # ── KPI Cards ────────────────────────────────────────
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 22

    kpis = [
        (5,  "B", "C", "💼  Valor total de compra (USD)",  f"='Mi Cartera'!F{fila_total}", "#,##0.00", AZUL_CLARO,   AZUL_OSCURO),
        (8,  "B", "C", "📈  Valor actual de la cartera",   f"='Mi Cartera'!G{fila_total}", "#,##0.00", VERDE_CLARO,  VERDE_OSCURO),
        (11, "B", "C", "💰  P&L total (USD)",              f"='Mi Cartera'!H{fila_total}", "+#,##0.00;-#,##0.00", AMARILLO, "8B6914"),
        (14, "B", "C", "📊  P&L total (%)",                f"='Mi Cartera'!I{fila_total}", '+0.00%;-0.00%',       AMARILLO, "8B6914"),
        (5,  "E", "F", "🏆  Mejor activo (mayor P&L %)",   f"=INDEX('Mi Cartera'!A{primera_data}:A{fila_total-1},MATCH(MAX('Mi Cartera'!I{primera_data}:I{fila_total-1}),'Mi Cartera'!I{primera_data}:I{fila_total-1},0))", "@", VERDE_CLARO, VERDE_OSCURO),
        (8,  "E", "F", "📉  Peor activo (menor P&L %)",    f"=INDEX('Mi Cartera'!A{primera_data}:A{fila_total-1},MATCH(MIN('Mi Cartera'!I{primera_data}:I{fila_total-1}),'Mi Cartera'!I{primera_data}:I{fila_total-1},0))", "@", ROJO_CLARO,  ROJO_OSCURO),
        (11, "E", "F", "🔢  Cantidad de activos",          f"=COUNTA('Mi Cartera'!A{primera_data}:A{fila_total-1})", "0", AZUL_CLARO, AZUL_OSCURO),
        (14, "E", "F", "📅  Última actualización",         "=TODAY()", "DD/MM/YYYY", GRIS_HEADER, "444444"),
    ]

    for fila, col_label, col_val, label, formula, fmt, color_f, color_t in kpis:
        # Label
        ws.row_dimensions[fila].height   = 22
        ws.row_dimensions[fila+1].height = 28

        ws.merge_cells(f"{col_label}{fila}:{col_val}{fila}")
        c = ws[f"{col_label}{fila}"]
        c.value     = label
        c.font      = Font(size=10, bold=True, color="444444", name="Calibri")
        c.fill      = fill(GRIS_HEADER)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = border_thin()

        # Valor
        ws.merge_cells(f"{col_label}{fila+1}:{col_val}{fila+1}")
        c = ws[f"{col_label}{fila+1}"]
        c.value         = formula
        c.number_format = fmt
        c.font          = Font(size=14, bold=True, color=color_t, name="Calibri")
        c.fill          = fill(color_f)
        c.alignment     = Alignment(horizontal="center", vertical="center")
        c.border        = border_medium()

    # ── Tabla de detalle por activo ───────────────────────
    fila_tabla = 18
    ws.merge_cells(f"B{fila_tabla}:F{fila_tabla}")
    c = ws[f"B{fila_tabla}"]
    c.value     = "DETALLE POR ACTIVO"
    c.font      = Font(bold=True, size=11, color=BLANCO, name="Calibri")
    c.fill      = fill(AZUL_OSCURO)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[fila_tabla].height = 26

    tabla_headers = ["Ticker", "P&L (USD)", "P&L (%)", "Valor actual", "Peso (%)"]
    cols_tabla    = ["B", "C", "D", "E", "F"]
    anchos_tabla  = [10, 16, 12, 16, 12]

    fila_th = fila_tabla + 1
    for col, header, ancho in zip(cols_tabla, tabla_headers, anchos_tabla):
        c = ws[f"{col}{fila_th}"]
        c.value = header
        estilo_header(c, color_fondo=AZUL_MED)
        ws.column_dimensions[col].width = ancho
    ws.row_dimensions[fila_th].height = 24

    for i in range(len(CARTERA_DEMO)):
        fila_ref = primera_data + i
        fila_d   = fila_th + 1 + i
        color_fila = GRIS_HEADER if i % 2 == 0 else BLANCO
        ws.row_dimensions[fila_d].height = 20

        formulas = [
            f"='Mi Cartera'!A{fila_ref}",
            f"='Mi Cartera'!H{fila_ref}",
            f"='Mi Cartera'!I{fila_ref}",
            f"='Mi Cartera'!G{fila_ref}",
            f"='Mi Cartera'!J{fila_ref}",
        ]
        formatos = ["@", "+#,##0.00;-#,##0.00", '+0.00%;-0.00%', "#,##0.00", '0.00"%"']

        for col, formula, fmt in zip(cols_tabla, formulas, formatos):
            c = ws[f"{col}{fila_d}"]
            c.value = formula
            estilo_celda(c, color_fondo=color_fila, alineacion="center", formato=fmt)

    print("  ✓ Hoja 'Resumen' creada")


# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*55}")
    print("  DASHBOARD DE CARTERA — FASE 3")
    print("  Generación del template Excel")
    print(f"{'='*55}\n")

    # Verificar input
    if not os.path.exists(INPUT_CSV):
        raise FileNotFoundError(
            f"No se encontró {INPUT_CSV}\n"
            "Ejecutá primero fase1_descarga_limpieza.py"
        )

    # Crear carpeta output si no existe
    os.makedirs("excel", exist_ok=True)

    # Cargar precios
    precios = pd.read_csv(INPUT_CSV, index_col=0, parse_dates=True)
    print(f"✓ CSV cargado: {len(precios)} días × {len(precios.columns)} activos\n")
    print("  Creando hojas del template...")
    print("  " + "─"*40)

    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja default vacía

    # Crear las 3 hojas
    primera_data, fila_total = crear_hoja_cartera(wb, precios)
    crear_hoja_precios(wb, precios)
    crear_hoja_resumen(wb, precios, primera_data, fila_total)

    # Guardar
    wb.save(OUTPUT_EXCEL)
    kb = os.path.getsize(OUTPUT_EXCEL) / 1024

    print(f"\n{'='*55}")
    print("  ✓ FASE 3 COMPLETADA")
    print(f"  Archivo generado:")
    print(f"    → {OUTPUT_EXCEL}  ({kb:.0f} KB)")
    print(f"  Hojas: Mi Cartera | Precios | Resumen")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
